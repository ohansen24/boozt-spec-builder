"""End-to-end `bsb run` on the real order files."""

from click.testing import CliRunner
from openpyxl import load_workbook
from tests.conftest import ODM_PATH, TEMPLATE_PATH

from bsb.cli import main


def test_bsb_run_end_to_end(tmp_path):
    out = tmp_path / "OR26BZQN0001_phase0.xlsx"
    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "run",
            "--odm",
            str(ODM_PATH),
            "--template",
            str(TEMPLATE_PATH),
            "--brand",
            "nars",
            "--out",
            str(out),
        ],
    )
    assert result.exit_code == 0, result.output
    assert out.exists()
    assert "119 rows" in result.output
    assert "Foundation" in result.output
    assert "Review queue" in result.output
    assert "pre-existing data rows" in result.output  # SVR leftovers warning

    wb = load_workbook(out)
    assert set(wb.sheetnames) >= {"Data sheet", "Provenance", "Run report"}


def test_bsb_run_unknown_brand(tmp_path):
    runner = CliRunner()
    result = runner.invoke(
        main,
        [
            "run",
            "--odm",
            str(ODM_PATH),
            "--template",
            str(TEMPLATE_PATH),
            "--brand",
            "gucci",
            "--out",
            str(tmp_path / "x.xlsx"),
        ],
    )
    assert result.exit_code != 0
    assert "unknown brand" in result.output


def test_resolve_rejects_brand_without_adapter():
    """No network: brands without an adapter must fail before any fetch."""
    runner = CliRunner()
    result = runner.invoke(main, ["resolve", "--gtin", "0194251147000", "--brand", "olaplex"])
    assert result.exit_code != 0
    assert "no adapter configured" in result.output
