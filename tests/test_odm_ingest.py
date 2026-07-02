"""ODM ingest (build kit 6.1) against the real OR26BZQN0001 order:
exactly 119 rows, all check digits valid, 27 base product groups."""

import pytest
from openpyxl import Workbook
from tests.conftest import ODM_PATH

from bsb.ingest.odm import (
    barcode_as_text,
    find_header_row,
    gs1_check_digit_ok,
    parse_odm,
    split_name,
)


@pytest.fixture(scope="module")
def odm():
    return parse_odm(ODM_PATH)


def test_exactly_119_rows(odm):
    assert len(odm.rows) == 119


def test_header_block_tolerated(odm):
    assert odm.header_row == 7  # metadata block sits above


def test_all_check_digits_valid_no_dupes(odm):
    assert odm.issues == []


def test_length_profile_all_upc12(odm):
    assert odm.length_profile == {12: 119}


def test_27_base_product_groups(odm):
    assert len(odm.base_names) == 27


def test_gtin13_is_zero_padded(odm):
    row = odm.rows[0]
    assert row.ean12 == "194251026404"
    assert row.gtin13 == "0194251026404"


def test_hints_extracted(odm):
    hints = odm.rows[0].hints
    assert hints["name"] == "Eyeshadow Quad - Orgasm"
    assert hints["coo"] == "CA"
    assert hints["qty"] == 12
    assert hints["price"] == 23.76
    assert hints["subcategory"] == "Eye Make-Up"
    assert hints["size"] == "4.4"
    assert hints["size_unit"] == "GR"


def test_shade_split(odm):
    assert odm.rows[0].base_name == "Eyeshadow Quad"
    assert odm.rows[0].shade == "Orgasm"
    assert split_name("Soft Matte Primer") == ("Soft Matte Primer", None)
    # only the FIRST " - " splits; en-dashes inside shades survive
    assert split_name("Foundation - New Shade L3.25 – Lima (Fn)") == (
        "Foundation",
        "New Shade L3.25 – Lima (Fn)",
    )


def test_barcodes_read_as_text_never_numbers():
    # leading zeros survive; floats with fractional parts are refused
    assert barcode_as_text("0123456789012") == "0123456789012"
    assert barcode_as_text(194251026404) == "194251026404"
    assert barcode_as_text(194251026404.0) == "194251026404"
    assert barcode_as_text(1942510264.5) is None
    assert barcode_as_text(None) is None
    assert barcode_as_text(" 194251026404\xa0") == "194251026404"


def test_gs1_check_digit():
    assert gs1_check_digit_ok("194251026404")  # real UPC from the order
    assert not gs1_check_digit_ok("194251026405")  # corrupted last digit
    assert gs1_check_digit_ok("0194251026404")  # 13-digit form of the same GTIN
    # the WIP paste-error EAN is a REAL SVR EAN — its check digit is valid,
    # which is exactly why lint catches it by EAN mismatch, not check digit
    assert gs1_check_digit_ok("3662361001699")
    assert not gs1_check_digit_ok("3662361001698")
    assert not gs1_check_digit_ok("abc")
    assert not gs1_check_digit_ok("1234")


def test_synthetic_leading_zero_barcode(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name", "QTY", "COO"])
    ws.append(["036000291452", "Test Product - Shade", 1, "US"])  # valid UPC, leading zero
    path = tmp_path / "odm.xlsx"
    wb.save(path)

    result = parse_odm(path)
    assert result.header_row == 1
    assert result.rows[0].ean12 == "036000291452"
    assert result.rows[0].gtin13 == "0036000291452"
    assert result.issues == []


def test_duplicate_detection(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name", "QTY"])
    ws.append(["194251026404", "A - X", 1])
    ws.append(["194251026404", "A - Y", 2])
    path = tmp_path / "odm.xlsx"
    wb.save(path)

    result = parse_odm(path)
    assert len(result.rows) == 2
    assert any("duplicate" in issue for issue in result.issues)


def test_missing_header_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Just", "Some", "Columns"])
    path = tmp_path / "odm.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="no header row"):
        parse_odm(path)


def test_find_header_row_requires_all_keys(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name"])  # QTY missing
    ws.append(["Barcode", "Name", "QTY"])
    path = tmp_path / "odm.xlsx"
    wb.save(path)
    assert find_header_row(wb.active) == 2


def test_zero_padded_ean13_flagged(tmp_path):
    """Review finding A: '0194251026404' passes GS1 mod-10 but is not a valid
    Boozt submission form — must surface as an ingest issue."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name", "QTY"])
    ws.append(["0194251026404", "Eyeshadow Quad - Orgasm", 1])
    path = tmp_path / "odm.xlsx"
    wb.save(path)

    result = parse_odm(path)
    assert any("not in a valid submission form" in i for i in result.issues)


def test_numeric_cell_dropping_leading_zero_flagged(tmp_path):
    """Review finding A: Excel storing UPC 036000291452 as a number yields an
    11-digit code whose GS1 check still passes — the length must be flagged."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name", "QTY"])
    ws.append([36000291452, "Test Product - Shade", 1])  # numeric cell, zero lost
    path = tmp_path / "odm.xlsx"
    wb.save(path)

    result = parse_odm(path)
    assert result.rows[0].ean12 == "36000291452"
    assert any("11 digits" in i and "36000291452" in i for i in result.issues)


def test_row_with_content_but_no_barcode_is_reported(tmp_path):
    """Review finding C: an ordered item must never vanish silently."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Barcode", "Name", "QTY"])
    ws.append(["194251026404", "A - X", 1])
    ws.append([None, "Concealer - Custard", 2])
    ws.append([None, None, None])  # genuine padding row stays silent
    path = tmp_path / "odm.xlsx"
    wb.save(path)

    result = parse_odm(path)
    assert len(result.rows) == 1
    content_issues = [i for i in result.issues if "no barcode but row has content" in i]
    assert len(content_issues) == 1
    assert "Concealer - Custard" in content_issues[0]
