"""Emit (build kit 6.9): colored template copy, EAN as text, cleared
pre-existing rows, Provenance + Run report sheets."""

import pytest
from openpyxl import load_workbook
from tests.conftest import ODM_PATH, TEMPLATE_PATH

from bsb.emit.writer import GREEN, RED, YELLOW, write_output
from bsb.ingest.odm import parse_odm
from bsb.ingest.template import map_headers
from bsb.pipeline import build_records


@pytest.fixture(scope="module")
def output(tmp_path_factory, brands, rules, synonyms):
    odm = parse_odm(ODM_PATH)
    records = build_records(odm, "nars", brands, rules, str(ODM_PATH))
    out = tmp_path_factory.mktemp("out") / "filled.xlsx"
    summary = write_output(
        TEMPLATE_PATH, out, records, synonyms, {"brand": "nars", "_ingest_issues": odm.issues}
    )
    wb = load_workbook(out)
    return summary, wb, records


def _fill_hex(cell):
    return cell.fill.start_color.rgb if cell.fill and cell.fill.fill_type == "solid" else None


def test_row_count_and_ean_text(output, synonyms):
    summary, wb, _records = output
    ws = wb["Data sheet"]
    tmap = map_headers(ws, synonyms)
    ean_col = tmap.columns["ean"]
    values = [ws.cell(row=r, column=ean_col).value for r in range(2, 2 + summary.records)]
    assert len(values) == 119
    assert values[0] == "194251026404"
    assert all(isinstance(v, str) for v in values)
    assert ws.cell(row=2, column=ean_col).number_format == "@"
    # nothing beyond the 119 rows
    assert ws.cell(row=2 + 119, column=ean_col).value is None


def test_preexisting_template_rows_cleared(output):
    summary, wb, _ = output
    # the "blank" template carried 52 leftover SVR rows
    assert summary.cleared_template_rows == 52
    ws = wb["Data sheet"]
    assert ws.cell(row=2, column=4).value != "SVR"


def test_fill_colors_by_status(output, synonyms):
    _summary, wb, _records = output
    ws = wb["Data sheet"]
    tmap = map_headers(ws, synonyms)
    row = 2  # Eyeshadow Quad - Orgasm

    def hex_of(field):
        return _fill_hex(ws.cell(row=row, column=tmap.columns[field]))

    green = GREEN.start_color.rgb
    yellow = YELLOW.start_color.rgb
    red = RED.start_color.rgb

    assert hex_of("country_iso") == green  # ODM_SOURCED
    assert hex_of("purchase_price") == green  # ODM_SOURCED
    assert hex_of("category") == yellow  # SINGLE_SOURCE rule
    assert hex_of("color_code") == yellow  # lexicon Orgasm -> 1003
    assert hex_of("style_number") == yellow  # MANUAL, empty, needs her input
    assert hex_of("style_name") == red  # NOT_FOUND (web source, Phase 1)
    assert hex_of("ingredients") == red
    assert hex_of("gender") is None  # MANUAL with value -> no fill
    assert hex_of("length") is None


def test_foundation_rows_get_yellow_1018(output, synonyms):
    _summary, wb, records = output
    ws = wb["Data sheet"]
    tmap = map_headers(ws, synonyms)
    cc_col = tmap.columns["color_code"]
    yellow = YELLOW.start_color.rgb

    foundation_rows = [2 + i for i, r in enumerate(records) if r.category.value == "Foundation"]
    assert len(foundation_rows) == 67
    for row in foundation_rows:
        cell = ws.cell(row=row, column=cc_col)
        assert cell.value == 1018
        assert _fill_hex(cell) == yellow


def test_values_written(output, synonyms):
    _summary, wb, _records = output
    ws = wb["Data sheet"]
    tmap = map_headers(ws, synonyms)
    row = 2
    assert ws.cell(row=row, column=tmap.columns["brand"]).value == "NARS"
    assert ws.cell(row=row, column=tmap.columns["gender"]).value == "F"
    assert ws.cell(row=row, column=tmap.columns["category"]).value == "Makeup"
    assert ws.cell(row=row, column=tmap.columns["country_iso"]).value == "CA"
    assert ws.cell(row=row, column=tmap.columns["length"]).value == "No Length"
    assert ws.cell(row=row, column=tmap.columns["variation"]).value == "No Variant"
    assert ws.cell(row=row, column=tmap.columns["purchase_price"]).value == 23.76
    assert ws.cell(row=row, column=tmap.columns["style_name"]).value is None  # never guessed
    assert ws.cell(row=row, column=tmap.columns["style_number"]).value is None  # open question 1


def test_provenance_sheet(output):
    _summary, wb, _records = output
    prov = wb["Provenance"]
    header = [c.value for c in prov[1]]
    assert header == [
        "ean",
        "field",
        "value",
        "status",
        "primary_url",
        "secondary_url",
        "method",
        "snippet",
        "notes",
    ]
    rows = list(prov.iter_rows(min_row=2, values_only=True))
    # every record contributes all 10 core fields + 4 extras
    assert len(rows) == 119 * 14
    coo_rows = [r for r in rows if r[0] == "194251026404" and r[1] == "country_iso"]
    assert coo_rows[0][3] == "ODM_SOURCED"
    assert coo_rows[0][6] == "odm"
    assert "COO=CA" in coo_rows[0][7]


def test_run_report_sheet_and_summary(output):
    summary, wb, _ = output
    assert "Run report" in wb.sheetnames
    assert summary.status_totals["ODM_SOURCED"] == 119 * 2  # country_iso + purchase_price
    assert summary.category_totals["Foundation"] == 67
    assert summary.category_totals["(uncategorized)"] == 1  # Light Reflecting Mist
    assert len(summary.review_red) > 0
    assert summary.unknown_headers  # e.g. Boozt Errors, How To Use
    # review queue is sorted red before yellow
    statuses = [i.status for i in summary.review_red] + [i.status for i in summary.review_yellow]
    first_yellow = next(
        (n for n, s in enumerate(statuses) if s in ("SINGLE_SOURCE", "MANUAL")), len(statuses)
    )
    assert all(s in ("CONFLICT", "NOT_FOUND") for s in statuses[:first_yellow])
