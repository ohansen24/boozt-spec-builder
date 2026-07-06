"""No-regression emit gate + field-wise merge (build kit 6.9).

Regression this locks in: a re-emit under a weaker pipeline config silently
overwrote a richer prior emit, dropping 42 sourced INCI cells. A re-emit must
never lose information — sourced value → empty, or green/yellow → red, fails
the emit unless explicitly allowed.
"""

import pytest
from tests.conftest import ODM_PATH, TEMPLATE_PATH

from bsb.emit.writer import RegressionError, _sanitize, detect_regressions, write_output
from bsb.ingest.odm import OdmRow, parse_odm
from bsb.models import FieldValue, ProductRecord
from bsb.pipeline import apply_retailer_primary, build_records
from bsb.resolve.generic import ResolverHit

# ---- detect_regressions (pure) -------------------------------------------

def test_detect_value_lost():
    prior = {("1", "ingredients"): ("Aqua, Glycerin", "SINGLE_SOURCE")}
    new = {("1", "ingredients"): (None, "NOT_FOUND")}
    assert detect_regressions(prior, new) == [
        "1 · ingredients: value 'Aqua, Glycerin' → (empty); status SINGLE_SOURCE → NOT_FOUND"
    ]


def test_detect_status_demoted_value_kept():
    # value still present but a green cell went red (e.g. new CONFLICT)
    prior = {("1", "style_name"): ("Foo", "VERIFIED")}
    new = {("1", "style_name"): ("Foo", "CONFLICT")}
    out = detect_regressions(prior, new)
    assert len(out) == 1 and "VERIFIED → CONFLICT" in out[0]


def test_detect_improvements_not_flagged():
    # empty→value and red→green are progress, never regressions
    prior = {
        ("1", "ingredients"): (None, "NOT_FOUND"),
        ("1", "style_name"): (None, "NOT_FOUND"),
        ("1", "size"): ("50 ml", "SINGLE_SOURCE"),  # unchanged
    }
    new = {
        ("1", "ingredients"): ("Aqua", "SINGLE_SOURCE"),
        ("1", "style_name"): ("Foo", "VERIFIED"),
        ("1", "size"): ("50 ml", "SINGLE_SOURCE"),
    }
    assert detect_regressions(prior, new) == []


def test_detect_dropped_row_is_regression():
    prior = {("1", "size"): ("50 ml", "SINGLE_SOURCE")}
    new = {}  # the ean vanished from the re-emit
    out = detect_regressions(prior, new)
    assert len(out) == 1 and "1 · size" in out[0]


def test_detect_by_design_blank_stable_not_flagged():
    # style_number is MANUAL/blank in both emits — no transition, no trip
    prior = {("1", "style_number"): (None, "MANUAL")}
    new = {("1", "style_number"): (None, "MANUAL")}
    assert detect_regressions(prior, new) == []


# ---- write_output gate (integration) -------------------------------------

@pytest.fixture
def base_records(brands, rules):
    # function-scoped: these tests mutate records, must not leak across tests
    odm = parse_odm(ODM_PATH)
    return build_records(odm, "nars", brands, rules, str(ODM_PATH))


def test_reemit_dropping_sourced_value_fails(tmp_path, base_records, synonyms):
    out = tmp_path / "order.xlsx"
    meta = {"brand": "nars", "_ingest_issues": []}
    # first emit: country_iso is ODM_SOURCED (green) for every row
    write_output(TEMPLATE_PATH, out, base_records, synonyms, dict(meta))
    assert out.exists()

    # second emit blanks a sourced field on one record -> must fail the gate
    victim = base_records[0]
    victim.country_iso = FieldValue(status="NOT_FOUND", notes="(simulated loss)")
    with pytest.raises(RegressionError) as ei:
        write_output(TEMPLATE_PATH, out, base_records, synonyms, dict(meta))
    assert any("country_iso" in line for line in ei.value.report)
    # the previous good emit was left intact (not overwritten by the failed run)
    from openpyxl import load_workbook

    wb = load_workbook(out)
    prov = wb["Provenance"]
    idx = {h: i for i, h in enumerate(next(prov.iter_rows(values_only=True)))}
    coo = [
        r for r in prov.iter_rows(min_row=2, values_only=True)
        if r[idx["ean"]] == victim.ean12 and r[idx["field"]] == "country_iso"
    ]
    assert coo and coo[0][idx["status"]] == "ODM_SOURCED"  # still the good value


def test_reemit_regression_allowed_with_flag(tmp_path, base_records, synonyms):
    out = tmp_path / "order.xlsx"
    meta = {"brand": "nars", "_ingest_issues": []}
    write_output(TEMPLATE_PATH, out, base_records, synonyms, dict(meta))
    base_records[0].country_iso = FieldValue(status="NOT_FOUND", notes="(intentional)")
    meta2 = dict(meta)
    summary = write_output(
        TEMPLATE_PATH, out, base_records, synonyms, meta2, allow_regressions=True
    )
    assert summary.records == len(base_records)
    # the allowance is recorded for the audit trail
    assert any("regressions allowed" in k for k in meta2)


def test_illegal_control_chars_are_stripped(tmp_path, base_records, synonyms):
    """Scraped INCI can carry control bytes openpyxl rejects — the writer must
    strip them (verbatim otherwise), never crash the emit."""
    assert _sanitize("Aqua,\x0b Glycerin\x0c, Parfum") == "Aqua, Glycerin, Parfum"
    assert _sanitize(None) is None
    victim = base_records[0]
    victim.ingredients = FieldValue(
        value="Aqua/Water/Eau, Glycerin\x0b, Parfum, Limonene", status="SINGLE_SOURCE"
    )
    out = tmp_path / "ctrl.xlsx"
    # must not raise IllegalCharacterError
    summary = write_output(TEMPLATE_PATH, out, base_records, synonyms, {"_ingest_issues": []})
    assert summary.records == len(base_records)
    from openpyxl import load_workbook

    prov = load_workbook(out)["Provenance"]
    idx = {h: i for i, h in enumerate(next(prov.iter_rows(values_only=True)))}
    vals = [
        r[idx["value"]]
        for r in prov.iter_rows(min_row=2, values_only=True)
        if r[idx["ean"]] == victim.ean12 and r[idx["field"]] == "ingredients"
    ]
    assert vals and "\x0b" not in vals[0]


def test_first_emit_has_no_prior_to_regress(tmp_path, base_records, synonyms):
    out = tmp_path / "fresh.xlsx"
    # no file at out yet -> gate is a no-op, emit succeeds
    summary = write_output(TEMPLATE_PATH, out, base_records, synonyms, {"_ingest_issues": []})
    assert summary.records == len(base_records)


# ---- field-wise merge: retailer-primary must not clobber -------------------

def test_retailer_primary_does_not_clobber_existing_inci(brands, rules):
    """apply_retailer_primary fills field-wise from evidence; a field it has no
    evidence for (here INCI) must keep whatever was already there — never
    last-write-wins a sourced value back to empty."""
    brand_cfg = brands["maria_nila"]
    row = OdmRow(
        row_number=1,
        ean12="7391681111111",
        gtin13="07391681111111",
        base_name="",
        shade=None,
        hints={},
    )
    record = ProductRecord(ean12=row.ean12, gtin13=row.gtin13, brand="Maria Nila")
    # pre-existing sourced INCI from a prior (e.g. brand-site) pass
    record.ingredients = FieldValue(value="Aqua, Glycerin", status="VERIFIED", notes="prior")

    # retailer hits carry a NAME but NO ingredients for this row
    hits = [
        ResolverHit(url="https://a.com/p", family="a.com", gtin_anchored=True, name="Shampoo"),
        ResolverHit(url="https://b.com/p", family="b.com", gtin_anchored=True, name="Shampoo"),
    ]
    apply_retailer_primary(record, row, hits, brand_cfg, rules)

    # name got filled from the two families; INCI was NOT touched
    assert record.style_name.value
    assert record.ingredients.value == "Aqua, Glycerin"
    assert record.ingredients.status == "VERIFIED"
