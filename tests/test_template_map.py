"""Header mapping (build kit 6.2) against all four real layouts — they
deliberately differ; that's the point. Never address columns by letter."""

from openpyxl import load_workbook
from tests.conftest import FIXTURES, TEMPLATE_PATH

from bsb.ingest.template import map_headers, normalize_header, read_sheet_rows


def _map(path, synonyms):
    ws = load_workbook(path)["Data sheet"]
    return map_headers(ws, synonyms)


def test_normalize_header():
    assert normalize_header("Gender (F = Female, M = Male, U = Unisex)") == "gender"
    assert normalize_header('Length (enter "No Length" if not applicable)') == "length"
    assert normalize_header("  Boozt\xa0Product   Category ") == "boozt product category"
    assert normalize_header("Swedish -  product description") == "swedish - product description"


def test_blank_template_layout(synonyms):
    tmap = _map(TEMPLATE_PATH, synonyms)
    assert tmap.columns["ean"] == 1
    assert tmap.columns["color_code"] == 11
    assert tmap.columns["flammable"] == 25
    assert tmap.columns["purchase_price"] == 32
    # this layout has no DG packing group / flash point columns
    assert "packing_group" in tmap.missing_fields
    assert "flash_point" in tmap.missing_fields
    unknown = [h for _, h in tmap.unknown_headers]
    assert "Boozt Errors" in unknown  # warned, left untouched


def test_nars_wip_layout(synonyms):
    tmap = _map(FIXTURES / "nars_wip.xlsx", synonyms)
    assert tmap.has("color_code")
    assert tmap.columns["purchase_price"] == 32


def test_olaplex_layout_has_no_color_code(synonyms):
    tmap = _map(FIXTURES / "olaplex_final.xlsx", synonyms)
    assert not tmap.has("color_code")  # skipped silently
    assert "color_code" in tmap.missing_fields
    assert tmap.columns["ingredients"] == 11  # Material composition sits at K here
    assert tmap.columns["country_iso"] == 15
    assert tmap.columns["purchase_price"] == 26


def test_aesop_layout_carries_dg_block(synonyms):
    tmap = _map(FIXTURES / "aesop_final.xlsx", synonyms)
    assert tmap.columns["packing_group"] == 29
    assert tmap.columns["flash_point"] == 31
    assert tmap.columns["country_iso"] == 25
    assert tmap.columns["purchase_price"] == 41


def test_header_variants_match_case_insensitively(synonyms):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["ean code", "STYLE/DISPLAY NAME", "Boozt  Color\xa0code", "Nonsense Column"])
    tmap = map_headers(ws, synonyms)
    assert tmap.columns == {"ean": 1, "style_name": 2, "color_code": 3}
    assert tmap.unknown_headers == [(4, "Nonsense Column")]


def test_read_sheet_rows_wip(synonyms):
    rows = read_sheet_rows(FIXTURES / "nars_wip.xlsx", synonyms)
    assert len(rows) == 119
    assert rows[0]["_row"] == 2
    assert rows[0]["style_number"] == "SVR3662361001699"
    assert rows[0]["size"] == "4,4gr"
