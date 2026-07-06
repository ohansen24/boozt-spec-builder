"""Output writer (build kit 6.9): a colored copy of the uploaded template plus
"Provenance" and "Run report" sheets.

The copy is header-mapped (never column letters), the EAN column is formatted
as text, and every pre-existing data row in the template is cleared first —
uploaded "blank" templates have been observed carrying leftover rows from
previous orders. Fill colors: green = VERIFIED / ODM_SOURCED, yellow =
SINGLE_SOURCE or needs-attention MANUAL, red = CONFLICT / NOT_FOUND.
"""

import shutil
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill
from pydantic import BaseModel, Field

from bsb.ingest.template import TemplateMap, map_headers
from bsb.models import FieldValue, ProductRecord
from bsb.validate.language import non_english_tokens

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

_STATUS_RANK = {"CONFLICT": 0, "NOT_FOUND": 0, "SINGLE_SOURCE": 1, "MANUAL": 1}

# no-regression guard (build kit 6.9): a re-emit must never silently lose
# information. "OK" = a value was present or the cell was green/yellow; "red" =
# the reviewer must source it.
_OK_STATUSES = {"VERIFIED", "ODM_SOURCED", "SINGLE_SOURCE", "MANUAL"}
_RED_STATUSES = {"NOT_FOUND", "CONFLICT"}


class RegressionError(RuntimeError):
    """A re-emit would drop information the previous emit had. Carries the
    per-cell report so the caller can print it and fail the run."""

    def __init__(self, report: list[str]):
        self.report = report
        super().__init__(
            f"emit would regress {len(report)} cell(s) vs the previous emit "
            "(sourced value lost or green/yellow → red)"
        )


def _cellval_empty(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _read_prior_cells(path: Path) -> dict[tuple[str, str], tuple[object, str]]:
    """(ean, field) -> (value, status) from a prior emit's Provenance sheet.
    Returns {} when the file/sheet is absent or unreadable — a first emit has
    nothing to regress against."""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    try:
        if "Provenance" not in wb.sheetnames:
            return {}
        prov = wb["Provenance"]
        rows = prov.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return {}
        idx = {str(h): i for i, h in enumerate(header)}
        need = ("ean", "field", "value", "status")
        if not all(k in idx for k in need):
            return {}
        out: dict[tuple[str, str], tuple[object, str]] = {}
        for r in rows:
            ean = r[idx["ean"]]
            field = r[idx["field"]]
            if ean is None or field is None:
                continue
            out[(str(ean), str(field))] = (r[idx["value"]], str(r[idx["status"]] or ""))
        return out
    finally:
        wb.close()


def _new_cells(records: list[ProductRecord]) -> dict[tuple[str, str], tuple[object, str]]:
    """(ean, field) -> (value, status) for the emit about to be written —
    mirrors what the Provenance sheet records (every field, template or not)."""
    out: dict[tuple[str, str], tuple[object, str]] = {}
    for record in records:
        fields = [(name, getattr(record, name)) for name in ProductRecord.field_values()]
        fields += list(record.extras.items())
        for field, fv in fields:
            out[(record.ean12, field)] = (fv.value, fv.status)
    return out


def detect_regressions(
    prior: dict[tuple[str, str], tuple[object, str]],
    new: dict[tuple[str, str], tuple[object, str]],
) -> list[str]:
    """Every cell that would go value→empty or OK(green/yellow)→red. Keyed by
    (ean, field) so it is insensitive to row reordering."""
    report: list[str] = []
    for key, (prior_val, prior_status) in sorted(prior.items()):
        new_val, new_status = new.get(key, (None, "MISSING"))
        ean, field = key
        lost_value = not _cellval_empty(prior_val) and _cellval_empty(new_val)
        demoted = prior_status in _OK_STATUSES and new_status in _RED_STATUSES
        if lost_value or demoted:
            reasons = []
            if lost_value:
                reasons.append(f"value {prior_val!r} → (empty)")
            if demoted:
                reasons.append(f"status {prior_status} → {new_status}")
            report.append(f"{ean} · {field}: " + "; ".join(reasons))
    return report


class ReviewItem(BaseModel):
    ean: str
    field: str
    status: str
    value: str | None = None
    notes: str = ""


class RunSummary(BaseModel):
    out_path: str
    records: int
    status_totals: dict[str, int] = Field(default_factory=dict)
    category_totals: dict[str, int] = Field(default_factory=dict)
    # NOT_FOUND split by failure class (Oli distinction): field -> count
    extraction_miss: dict[str, int] = Field(default_factory=dict)
    no_source: dict[str, int] = Field(default_factory=dict)
    # QA: shipped style_name/color_name values that read as non-English
    non_english_names: list[ReviewItem] = Field(default_factory=list)
    review_red: list[ReviewItem] = Field(default_factory=list)
    review_yellow: list[ReviewItem] = Field(default_factory=list)
    cleared_template_rows: int = 0
    unknown_headers: list[str] = Field(default_factory=list)
    missing_fields: list[str] = Field(default_factory=list)
    ingest_issues: list[str] = Field(default_factory=list)
    verify_at_receipt: list[ReviewItem] = Field(default_factory=list)


def _by_design_blank(fv: FieldValue) -> bool:
    """MANUAL blanks whose notes start with "by design:" are intentional
    (e.g. style number issued by Boozt after receipt) — no fill, no review."""
    return fv.status == "MANUAL" and not fv.value and fv.notes.startswith("by design:")


def fill_for(fv: FieldValue) -> PatternFill | None:
    if fv.status in ("VERIFIED", "ODM_SOURCED"):
        return GREEN
    if fv.status == "SINGLE_SOURCE":
        return YELLOW
    if fv.status in ("CONFLICT", "NOT_FOUND"):
        return RED
    if fv.status == "MANUAL":
        return None if fv.value or _by_design_blank(fv) else YELLOW
    return None


def _cell_value(field: str, fv: FieldValue) -> object:
    if fv.value is None:
        return None
    if field == "color_code":
        try:
            return int(fv.value)
        except ValueError:
            return _sanitize(fv.value)
    if field == "purchase_price":
        try:
            return float(fv.value)
        except ValueError:
            return _sanitize(fv.value)
    return _sanitize(fv.value)


def _sanitize(value: object) -> object:
    """Strip control characters Excel/openpyxl rejects (IllegalCharacterError).
    Scraped page text (notably INCI copied from PDPs) can carry stray control
    bytes; the value is otherwise kept verbatim (tabs/newlines preserved)."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _clear_data_rows(ws, tmap: TemplateMap) -> int:
    """Remove pre-existing data rows below the header from the output copy."""
    ean_col = tmap.columns.get("ean")
    cleared = 0
    if ean_col is not None:
        for row in ws.iter_rows(min_row=tmap.header_row + 1):
            if row[ean_col - 1].value not in (None, ""):
                cleared += 1
    if ws.max_row > tmap.header_row:
        ws.delete_rows(tmap.header_row + 1, ws.max_row - tmap.header_row)
    return cleared


def write_output(
    template_path: str | Path,
    out_path: str | Path,
    records: list[ProductRecord],
    synonyms: dict[str, list[str]],
    run_meta: dict,
    allow_regressions: bool = False,
) -> RunSummary:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # no-regression gate: compare against the previous emit at this path BEFORE
    # it is overwritten. A re-emit that would drop a sourced value or demote a
    # green/yellow cell to red fails the run (unless explicitly allowed) so
    # information is never silently lost across runs.
    regressions = detect_regressions(_read_prior_cells(out_path), _new_cells(records))
    if regressions and not allow_regressions:
        raise RegressionError(regressions)
    if regressions:
        run_meta["regressions allowed (--allow-regressions)"] = (
            f"{len(regressions)} cell(s) lost/demoted vs previous emit: "
            + " | ".join(regressions[:20])
            + (" …" if len(regressions) > 20 else "")
        )

    shutil.copyfile(template_path, out_path)

    wb = load_workbook(out_path)
    ws = wb.worksheets[0]
    tmap = map_headers(ws, synonyms)
    if not tmap.has("ean"):
        raise ValueError(f"{template_path}: no EAN column found via header synonyms")

    cleared = _clear_data_rows(ws, tmap)

    non_english: list[ReviewItem] = []  # QA: shipped names that read non-English
    status_counter: Counter[str] = Counter()
    # Oli distinction: a red cell backed by an anchored source (has a primary
    # SourceRef) is an EXTRACTION MISS — we reached the product page but pulled
    # nothing (an extractor/source-selection problem). A red cell with no
    # source is NO SOURCE FOUND (a discovery problem). Counted separately so
    # the two failure classes never blur together in a run report.
    anchored_miss: Counter[str] = Counter()
    no_source: Counter[str] = Counter()
    category_counter: Counter[str] = Counter()
    review: list[tuple[int, ReviewItem]] = []
    provenance_rows: list[list[object]] = []

    for i, record in enumerate(records):
        row_no = tmap.header_row + 1 + i
        ean_cell = ws.cell(row=row_no, column=tmap.columns["ean"])
        ean_cell.value = record.ean12
        ean_cell.number_format = "@"

        if tmap.has("brand"):
            ws.cell(row=row_no, column=tmap.columns["brand"]).value = record.brand

        fields: list[tuple[str, FieldValue]] = [
            (name, getattr(record, name)) for name in ProductRecord.field_values()
        ]
        fields += list(record.extras.items())

        for field, fv in fields:
            in_template = tmap.has(field)
            if in_template:
                cell = ws.cell(row=row_no, column=tmap.columns[field])
                cell.value = _cell_value(field, fv)
                fill = fill_for(fv)
                if fill is not None:
                    cell.fill = fill

                status_counter[fv.status] += 1
                if fv.status == "NOT_FOUND":
                    (anchored_miss if fv.primary is not None else no_source)[field] += 1
                # QA sweep: a shipped English-required name must read English —
                # a human caught 2 non-English names once; the machine catches
                # the next (Boozt requires English style_name/color_name)
                if field in ("style_name", "color_name") and fv.value:
                    toks = non_english_tokens(str(fv.value))
                    if toks:
                        non_english.append(
                            ReviewItem(
                                ean=record.ean12,
                                field=field,
                                status=fv.status,
                                value=fv.value,
                                notes=f"non-English tokens: {', '.join(toks)}",
                            )
                        )
                rank = _STATUS_RANK.get(fv.status)
                settled_manual = fv.status == "MANUAL" and (fv.value or _by_design_blank(fv))
                if rank is not None and not settled_manual:
                    review.append(
                        (
                            rank,
                            ReviewItem(
                                ean=record.ean12,
                                field=field,
                                status=fv.status,
                                value=fv.value,
                                notes=fv.notes,
                            ),
                        )
                    )

            provenance_rows.append(
                [
                    record.ean12,
                    field,
                    fv.value,
                    fv.status,
                    fv.primary.url if fv.primary else None,
                    fv.secondary.url if fv.secondary else None,
                    fv.primary.method if fv.primary else None,
                    fv.primary.snippet if fv.primary else None,
                    fv.notes
                    if in_template
                    else f"{fv.notes} [column absent from template]".strip(),
                ]
            )

        if record.category.value:
            category_counter[record.category.value] += 1
        else:
            category_counter["(uncategorized)"] += 1

    prov = wb.create_sheet("Provenance")
    prov.append(
        [
            "ean",
            "field",
            "value",
            "status",
            "primary_url",
            "secondary_url",
            "method",
            "snippet",
            "notes",
        ]
    )
    for row in provenance_rows:
        prov.append([_sanitize(c) for c in row])
    for cell in prov["A"]:
        cell.number_format = "@"

    review.sort(key=lambda item: (item[0], item[1].ean, item[1].field))
    review_red = [r for rank, r in review if rank == 0]
    review_yellow = [r for rank, r in review if rank == 1]

    # receiving checklist: cells flagged for confirmation against physical
    # goods at warehouse receipt (Felina's checkpoint), independent of color
    receipt_checks = [
        ReviewItem(ean=r.ean12, field=f, status=fv.status, value=fv.value, notes=fv.notes)
        for r in records
        for f, fv in (
            [(name, getattr(r, name)) for name in ProductRecord.field_values()]
            + list(r.extras.items())
        )
        if "VERIFY_AT_RECEIPT" in fv.notes
    ]

    report = wb.create_sheet("Run report")
    for key, value in run_meta.items():
        if key.startswith("_"):
            continue
        report.append([key, str(value)])
    report.append(["records", len(records)])
    report.append(["cleared pre-existing template data rows", cleared])
    if tmap.unknown_headers:
        report.append(
            ["unknown headers (left untouched)", "; ".join(h for _, h in tmap.unknown_headers)]
        )
    if tmap.missing_fields:
        report.append(
            ["canonical fields absent from template (skipped)", "; ".join(tmap.missing_fields)]
        )
    for issue in run_meta.get("_ingest_issues", []):
        report.append(["ingest issue", issue])
    report.append([])
    report.append(["status", "cells"])
    for status, count in sorted(status_counter.items()):
        report.append([status, count])
    if non_english:
        report.append([])
        report.append(["QA — non-English tokens in a name (Boozt requires English)"])
        report.append(["ean", "field", "value", "tokens"])
        for item in non_english:
            report.append([item.ean, item.field, _sanitize(item.value), item.notes])
    report.append([])
    report.append(["red cells by failure class", "cells"])
    report.append(
        ["extraction miss (anchored source, nothing extracted)", sum(anchored_miss.values())]
    )
    for field, count in sorted(anchored_miss.items(), key=lambda kv: -kv[1]):
        report.append([f"  extraction miss · {field}", count])
    report.append(["no source found (no anchored page)", sum(no_source.values())])
    for field, count in sorted(no_source.items(), key=lambda kv: -kv[1]):
        report.append([f"  no source · {field}", count])
    report.append([])
    report.append(["category", "rows"])
    for category, count in sorted(category_counter.items()):
        report.append([category, count])
    if receipt_checks:
        report.append([])
        report.append(["VERIFY AT RECEIPT — confirm against physical goods (warehouse)"])
        report.append(["ean", "field", "value", "notes"])
        for item in receipt_checks:
            report.append([item.ean, item.field, _sanitize(item.value), _sanitize(item.notes)])
    report.append([])
    report.append(["review queue (red first, then yellow)"])
    report.append(["ean", "field", "status", "value", "notes"])
    for item in review_red + review_yellow:
        report.append(
            [item.ean, item.field, item.status, _sanitize(item.value), _sanitize(item.notes)]
        )

    wb.save(out_path)

    return RunSummary(
        out_path=str(out_path),
        records=len(records),
        status_totals=dict(status_counter),
        category_totals=dict(category_counter),
        extraction_miss=dict(anchored_miss),
        no_source=dict(no_source),
        non_english_names=non_english,
        review_red=review_red,
        review_yellow=review_yellow,
        cleared_template_rows=cleared,
        unknown_headers=[h for _, h in tmap.unknown_headers],
        missing_fields=tmap.missing_fields,
        ingest_issues=list(run_meta.get("_ingest_issues", [])),
        verify_at_receipt=receipt_checks,
    )
