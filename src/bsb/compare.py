"""Compare an externally hand-filled sheet against the tool's output.

Both sides are normalized before comparison (NBSP/whitespace, case, decimal
commas; sizes through the guide normalizer, INCI through the token compare),
so pure formatting noise never shows up as a value difference. Every real
difference lands in a session workbook with both values, the tool's
provenance URL, and a blank Decision column — nothing is ingested from the
external sheet until each DISAGREE row carries a decision.
"""

import re

from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from bsb.ingest.template import read_sheet_rows
from bsb.normalize.boozt import clean_ws, normalize_size
from bsb.validate.matrix import compare_inci

_DECIMAL_COMMA = re.compile(r"(\d),(\d)")

CLASS_ORDER = {"DISAGREE": 0, "FORMAT_ONLY": 1, "FELINA_ONLY": 2, "TOOL_ONLY": 3}


class Difference(BaseModel):
    ean: str
    field: str
    classification: str  # DISAGREE | TOOL_ONLY | FELINA_ONLY
    ours: str | None = None
    theirs: str | None = None
    provenance_url: str | None = None
    note: str = ""


class ComparisonResult(BaseModel):
    compared_cells: int = 0
    agree: int = 0
    differences: list[Difference] = Field(default_factory=list)
    theirs_whitespace_dirty: int = 0
    fields: list[str] = Field(default_factory=list)
    eans_ours_only: list[str] = Field(default_factory=list)
    eans_theirs_only: list[str] = Field(default_factory=list)

    @property
    def disagreements(self) -> list[Difference]:
        return [d for d in self.differences if d.classification == "DISAGREE"]


def _base_norm(value: object) -> str | None:
    """Whitespace/decimal normalization only (case- and accent-sensitive)."""
    if value is None:
        return None
    text = clean_ws(str(value))
    text = " ".join(text.split())
    text = _DECIMAL_COMMA.sub(r"\1.\2", text)
    return text or None


def norm_text(value: object) -> str | None:
    """Full comparison form: base normalization + casefold. Never displayed."""
    base = _base_norm(value)
    return base.casefold() if base is not None else None


def _fold(value: str) -> str:
    import unicodedata

    return "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )


def _classify_values(field: str, ours: object, theirs: object) -> tuple[str, str]:
    """Field-aware comparison -> (classification, note). AGREE covers pure
    whitespace/decimal noise; FORMAT_ONLY covers case- or diacritics-only
    differences (surfaced, but not decisions); DISAGREE is a real difference."""
    if field == "ingredients":
        verdict, diff = compare_inci(str(ours), str(theirs))
        if verdict == "identical":
            return "AGREE", ""
        return "DISAGREE", f"INCI {verdict}: {diff}"[:160]
    if field == "size":
        a = normalize_size(ours) or norm_text(ours)
        b = normalize_size(theirs) or norm_text(theirs)
        return ("AGREE" if a == b else "DISAGREE"), ""
    if field == "color_code":
        try:
            # spreadsheet cells arrive as int, float, or text
            same = int(float(str(ours).strip())) == int(float(str(theirs).strip()))
            return ("AGREE" if same else "DISAGREE"), ""
        except ValueError:
            pass

    a_base, b_base = _base_norm(ours), _base_norm(theirs)
    if a_base == b_base:
        return "AGREE", ""
    a_case, b_case = (a_base or "").casefold(), (b_base or "").casefold()
    if a_case == b_case:
        return "FORMAT_ONLY", "case-only difference"
    if _fold(a_case) == _fold(b_case):
        return "FORMAT_ONLY", "diacritics-only difference"
    return "DISAGREE", ""


def _ean_key(value: object) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits or None


def load_provenance_urls(ours_path: str) -> dict[tuple[str, str], str]:
    wb = load_workbook(ours_path, read_only=True)
    if "Provenance" not in wb.sheetnames:
        return {}
    urls: dict[tuple[str, str], str] = {}
    for row in wb["Provenance"].iter_rows(min_row=2, values_only=True):
        ean, field, url = str(row[0] or ""), str(row[1] or ""), row[4]
        if url:
            urls[(ean, field)] = str(url)
    return urls


def compare_sheets(ours_path: str, theirs_path: str, synonyms: dict) -> ComparisonResult:
    ours_rows = {
        r: row for row in read_sheet_rows(ours_path, synonyms) if (r := _ean_key(row.get("ean")))
    }
    theirs_rows = {
        r: row for row in read_sheet_rows(theirs_path, synonyms) if (r := _ean_key(row.get("ean")))
    }
    provenance = load_provenance_urls(ours_path)

    shared_fields = sorted(
        (set(next(iter(ours_rows.values()))) & set(next(iter(theirs_rows.values()))))
        - {"_row", "ean"}
    )

    result = ComparisonResult(fields=shared_fields)
    result.eans_ours_only = sorted(set(ours_rows) - set(theirs_rows))
    result.eans_theirs_only = sorted(set(theirs_rows) - set(ours_rows))

    # whitespace hygiene on their side (count only; normalization hides these
    # from the value diff by design)
    for row in theirs_rows.values():
        for field, value in row.items():
            if field != "_row" and isinstance(value, str) and value != clean_ws(value):
                result.theirs_whitespace_dirty += 1

    for ean in ours_rows:
        if ean not in theirs_rows:
            continue
        ours_row, theirs_row = ours_rows[ean], theirs_rows[ean]
        for field in shared_fields:
            a, b = ours_row.get(field), theirs_row.get(field)
            a_empty = a in (None, "") or (isinstance(a, str) and not a.strip())
            b_empty = b in (None, "") or (isinstance(b, str) and not b.strip())
            if a_empty and b_empty:
                continue
            result.compared_cells += 1
            url = provenance.get((ean, field))
            if a_empty:
                result.differences.append(
                    Difference(
                        ean=ean,
                        field=field,
                        classification="FELINA_ONLY",
                        theirs=str(b),
                        provenance_url=url,
                    )
                )
                continue
            if b_empty:
                result.differences.append(
                    Difference(
                        ean=ean,
                        field=field,
                        classification="TOOL_ONLY",
                        ours=str(a),
                        provenance_url=url,
                    )
                )
                continue
            classification, note = _classify_values(field, a, b)
            if classification == "AGREE":
                result.agree += 1
            else:
                result.differences.append(
                    Difference(
                        ean=ean,
                        field=field,
                        classification=classification,
                        ours=str(a),
                        theirs=str(b),
                        provenance_url=url,
                        note=note,
                    )
                )

    result.differences.sort(key=lambda d: (CLASS_ORDER[d.classification], d.field, d.ean))
    return result


def write_comparison(
    result: ComparisonResult, out_path: str, ours_path: str, theirs_path: str
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    pct = (result.agree / result.compared_cells * 100) if result.compared_cells else 0.0
    summary.append(
        ["AGREEMENT", f"{result.agree} of {result.compared_cells} cells identical ({pct:.1f}%)"]
    )
    summary.append(["tool sheet", ours_path])
    summary.append(["external sheet", theirs_path])
    summary.append([])
    counts = {}
    for d in result.differences:
        counts[d.classification] = counts.get(d.classification, 0) + 1
    for cls in ("DISAGREE", "FORMAT_ONLY", "FELINA_ONLY", "TOOL_ONLY"):
        summary.append([cls, counts.get(cls, 0)])
    summary.append(
        [
            "whitespace-dirty cells in external sheet (normalized away)",
            result.theirs_whitespace_dirty,
        ]
    )
    if result.eans_ours_only:
        summary.append(["EANs only in tool sheet", "; ".join(result.eans_ours_only)])
    if result.eans_theirs_only:
        summary.append(["EANs only in external sheet", "; ".join(result.eans_theirs_only)])
    summary.append([])
    summary.append(["Nothing from the external sheet becomes config until every"])
    summary.append(["DISAGREE row below carries a Decision."])

    diff_ws = wb.create_sheet("Differences")
    diff_ws.append(
        [
            "ean",
            "field",
            "classification",
            "tool value",
            "felina value",
            "tool provenance",
            "note",
            "Decision",
        ]
    )
    for d in result.differences:
        diff_ws.append(
            [d.ean, d.field, d.classification, d.ours, d.theirs, d.provenance_url, d.note, None]
        )
    for cell in diff_ws["A"]:
        cell.number_format = "@"
    wb.save(out_path)
