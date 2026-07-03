"""bsb ingest-portal-errors: the Boozt portal is the final QA stage — its
rejections must compound like everything else. Reads the "Boozt Errors"
column of an uploaded-and-returned sheet (or free-text notes), classifies
each error to the field it most likely concerns, and drafts order-override
entries for human review. Nothing is applied automatically: the draft lands
next to the order's override file and is merged by hand after review.
"""

import re
from pathlib import Path

from pydantic import BaseModel, Field

from bsb.ingest.template import map_headers

# error text -> the field it most likely concerns (first match wins)
_FIELD_HINTS: list[tuple[str, str]] = [
    (r"colou?r\s*code|farbcode|\b10[0-2][0-9]\b", "color_code"),
    (r"colou?r|shade|farbe", "color_name"),
    (r"size|volume|ml\b|gramm|\bg\b", "size"),
    (r"categor", "category"),
    (r"ingredient|inci|material composition", "ingredients"),
    (r"ean|gtin|barcode", "ean"),
    (r"style\s*number", "style_number"),
    (r"style|display\s*name|product\s*name", "style_name"),
    (r"gender", "gender"),
    (r"flammable|hazard|\bun\b|dangerous", "flammable"),
    (r"country|origin|iso", "country_iso"),
]


class PortalError(BaseModel):
    ean: str
    error: str
    field_guess: str = "unknown"


class PortalIngest(BaseModel):
    order: str
    errors: list[PortalError] = Field(default_factory=list)
    draft_path: str | None = None

    def by_field(self) -> dict[str, int]:
        counts: dict[str, int] = {}
        for e in self.errors:
            counts[e.field_guess] = counts.get(e.field_guess, 0) + 1
        return counts


def _guess_field(error_text: str) -> str:
    lowered = error_text.casefold()
    for pattern, field in _FIELD_HINTS:
        if re.search(pattern, lowered):
            return field
    return "unknown"


def collect_portal_errors(sheet_path: str, synonyms: dict) -> list[PortalError]:
    """Rows with a non-empty 'Boozt Errors' cell (the column is not in the
    canonical synonym table by design — it is portal-owned)."""
    from openpyxl import load_workbook

    wb = load_workbook(sheet_path, data_only=True)
    ws = wb.worksheets[0]
    tmap = map_headers(ws, synonyms)
    error_col = next(
        (col for col, header in tmap.unknown_headers if "boozt errors" in header.casefold()),
        None,
    )
    if error_col is None:
        raise ValueError(f"{sheet_path}: no 'Boozt Errors' column found")
    ean_col = tmap.columns["ean"]

    errors: list[PortalError] = []
    for row in ws.iter_rows(min_row=tmap.header_row + 1):
        ean = str(row[ean_col - 1].value or "").strip()
        error = str(row[error_col - 1].value or "").strip()
        if ean and error:
            errors.append(PortalError(ean=ean, error=error, field_guess=_guess_field(error)))
    return errors


def draft_overrides(order: str, errors: list[PortalError], out_dir: Path) -> Path:
    """Draft override entries — value left blank on purpose: the portal says
    what is WRONG; a human decides what is right."""
    lines = [
        f"# DRAFT from portal errors ({order}) — review, fill `value`, then",
        f"# merge the entries you accept into config/order_overrides/{order}.yaml.",
        "# Recurring patterns belong in rules/lexicon config instead, so the",
        "# fix compounds beyond this order.",
        f"order: {order}",
        "overrides:",
    ]
    for e in errors:
        lines += [
            f"  - field: {e.field_guess if e.field_guess != 'unknown' else 'FIXME'}",
            f'    eans: ["{e.ean}"]',
            "    value: FIXME",
            "    status: VERIFIED",
            "    decided_by: portal+human",
            "    date: FIXME",
            f"    rationale: 'portal rejection: {e.error[:160]}'",
        ]
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{order}_portal_draft.yaml"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
