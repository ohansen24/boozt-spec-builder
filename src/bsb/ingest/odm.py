"""ODM parser, header-block tolerant (build kit section 6.1).

The table header row is located by finding the row containing at least
{"Barcode", "Name", "QTY"}; a metadata block may sit above it. Barcodes are
read as text, never as numbers (leading-zero and precision safety). Ingest
checks: GS1 mod-10 check digit on every barcode, duplicate detection, and a
length profile.
"""

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, Field

HEADER_REQUIRED = {"Barcode", "Name", "QTY"}

# ODM header text -> hint key (build kit 6.1: Name, Size, Size Unit, COO,
# Gender, Client Price, QTY, Subcategory).
HINT_COLUMNS = {
    "Name": "name",
    "Size": "size",
    "Size Unit": "size_unit",
    "COO": "coo",
    "Gender": "gender",
    "Client Price": "price",
    "QTY": "qty",
    "Subcategory": "subcategory",
    "Category": "category",
}


class OdmRow(BaseModel):
    row_number: int  # 1-based worksheet row, for error messages
    ean12: str  # barcode exactly as in the ODM (12-digit UPC in this order)
    gtin13: str  # "0" + ean12 when 12 digits, used for site lookups
    base_name: str  # product name without shade ("Eyeshadow Quad")
    shade: str | None  # shade part after " - " ("Orgasm"), None if absent
    hints: dict = Field(default_factory=dict)


class OdmParseResult(BaseModel):
    rows: list[OdmRow]
    header_row: int
    issues: list[str] = Field(default_factory=list)
    length_profile: dict[int, int] = Field(default_factory=dict)

    @property
    def base_names(self) -> list[str]:
        seen: dict[str, None] = {}
        for r in self.rows:
            seen.setdefault(r.base_name)
        return list(seen)


def gs1_check_digit_ok(code: str) -> bool:
    """GS1 mod-10 check digit, valid for GTIN-8/12/13/14."""
    if not code.isdigit() or len(code) < 8:
        return False
    digits = [int(c) for c in code]
    body = digits[:-1][::-1]
    total = sum(d * (3 if i % 2 == 0 else 1) for i, d in enumerate(body))
    return (10 - total % 10) % 10 == digits[-1]


def check_ean_submission_form(ean: str) -> bool:
    """Boozt: an EAN-13 must not start with 0 — 12-digit UPCs are submitted
    as-is, never zero-padded to 13. Also refuses stray lengths (e.g. the
    11-digit code left when Excel stores a 0-leading UPC as a number)."""
    if not ean.isdigit():
        return False
    if len(ean) == 13 and ean.startswith("0"):
        return False
    return len(ean) in (8, 12, 13)


def barcode_as_text(value: object) -> str | None:
    """Coerce a barcode cell to its text form without losing leading zeros.

    Cells stored as numbers are recovered via int(); floats with a fractional
    part are refused rather than silently rounded.
    """
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip().replace("\xa0", "")
        return text or None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not value.is_integer():
            return None
        return str(int(value))
    return None


def split_name(name: str) -> tuple[str, str | None]:
    """Split an ODM product name into (base, shade) on the first " - "."""
    base, sep, shade = name.partition(" - ")
    if not sep:
        return name.strip(), None
    return base.strip(), shade.strip()


def find_header_row(ws) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=50):
        values = {str(c.value).strip() for c in row if c.value not in (None, "")}
        if values >= HEADER_REQUIRED:
            return row[0].row
    return None


def parse_odm(path: str | Path) -> OdmParseResult:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header_row = find_header_row(ws)
    if header_row is None:
        raise ValueError(
            f"{path}: no header row containing {sorted(HEADER_REQUIRED)} in the first 50 rows"
        )

    col_by_header: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value not in (None, ""):
            col_by_header.setdefault(str(cell.value).strip(), cell.column)
    barcode_col = col_by_header["Barcode"]

    rows: list[OdmRow] = []
    issues: list[str] = []
    seen: dict[str, int] = {}

    content_cols = [col_by_header[h] for h in ("Name", "QTY") if h in col_by_header]

    for row in ws.iter_rows(min_row=header_row + 1):
        raw = row[barcode_col - 1].value
        row_number = row[0].row
        if raw in (None, ""):
            # blank padding rows are fine, but a row with content and no
            # barcode is a vanishing ordered item — never drop it silently
            content = [row[c - 1].value for c in content_cols]
            if any(v not in (None, "") for v in content):
                issues.append(
                    f"row {row_number}: no barcode but row has content "
                    f"({content[0]!r}) — row skipped, resolve in the ODM"
                )
            continue
        barcode = barcode_as_text(raw)
        if barcode is None or not barcode.isdigit():
            issues.append(f"row {row_number}: unreadable barcode {raw!r}")
            continue

        if not gs1_check_digit_ok(barcode):
            issues.append(f"row {row_number}: GS1 check digit failed for {barcode}")
        if not check_ean_submission_form(barcode):
            issues.append(
                f"row {row_number}: EAN {barcode} is not in a valid submission form "
                f"({len(barcode)} digits; an EAN-13 must not start with 0, "
                "12-digit UPCs are submitted as-is)"
            )
        if barcode in seen:
            issues.append(
                f"row {row_number}: duplicate barcode {barcode} (first at row {seen[barcode]})"
            )
        else:
            seen[barcode] = row_number

        hints = {}
        for header, key in HINT_COLUMNS.items():
            col = col_by_header.get(header)
            if col is None:
                continue
            value = row[col - 1].value
            if isinstance(value, str):
                value = value.strip()
            hints[key] = value

        name = str(hints.get("name") or "")
        base, shade = split_name(name)
        rows.append(
            OdmRow(
                row_number=row_number,
                ean12=barcode,
                gtin13="0" + barcode if len(barcode) == 12 else barcode,
                base_name=base,
                shade=shade,
                hints=hints,
            )
        )

    return OdmParseResult(
        rows=rows,
        header_row=header_row,
        issues=issues,
        length_profile=dict(Counter(len(r.ean12) for r in rows)),
    )
