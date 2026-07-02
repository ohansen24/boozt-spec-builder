"""Template header mapping through config/header_synonyms.yaml (build kit 6.2).

Never address columns by letter: the three finished sheets do not share one
layout. Headers are matched case-insensitively on a normalized form (NBSP to
space, parentheticals stripped, whitespace collapsed). Unknown headers are
reported and left untouched; canonical fields absent from a template are
skipped silently (the Olaplex layout has no color_code).
"""

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, Field

_PARENTHETICAL = re.compile(r"\([^()]*\)")
_WHITESPACE = re.compile(r"\s+")


def normalize_header(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = _PARENTHETICAL.sub(" ", text)
    text = _WHITESPACE.sub(" ", text)
    return text.strip().casefold()


class TemplateMap(BaseModel):
    header_row: int = 1
    columns: dict[str, int] = Field(default_factory=dict)  # canonical -> 1-based column
    unknown_headers: list[tuple[int, str]] = Field(default_factory=list)
    missing_fields: list[str] = Field(default_factory=list)

    def has(self, canonical: str) -> bool:
        return canonical in self.columns


def build_synonym_index(synonyms: dict[str, list[str]]) -> dict[str, str]:
    """normalized observed header -> canonical name."""
    index: dict[str, str] = {}
    for canonical, observed_list in synonyms.items():
        for observed in observed_list:
            key = normalize_header(observed)
            existing = index.get(key)
            if existing is not None and existing != canonical:
                raise ValueError(
                    f"header synonym {observed!r} maps to both {existing!r} and {canonical!r}"
                )
            index[key] = canonical
    return index


def map_headers(ws: Worksheet, synonyms: dict[str, list[str]], header_row: int = 1) -> TemplateMap:
    """Map row `header_row` of a template sheet to canonical field names."""
    index = build_synonym_index(synonyms)
    columns: dict[str, int] = {}
    unknown: list[tuple[int, str]] = []

    for cell in ws[header_row]:
        if cell.value in (None, ""):
            continue
        canonical = index.get(normalize_header(str(cell.value)))
        if canonical is None:
            unknown.append((cell.column, str(cell.value)))
        elif canonical not in columns:  # first occurrence wins
            columns[canonical] = cell.column

    missing = [c for c in synonyms if c not in columns]
    return TemplateMap(
        header_row=header_row, columns=columns, unknown_headers=unknown, missing_fields=missing
    )


def read_sheet_rows(
    path: str | Path, synonyms: dict[str, list[str]], sheet: str | None = None
) -> list[dict[str, object]]:
    """Read a filled data sheet into canonical-keyed dicts (raw cell values).

    Used for linting existing sheets and for golden replays. Rows without an
    EAN are skipped. Each dict carries "_row" with the worksheet row number.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    tmap = map_headers(ws, synonyms)
    ean_col = tmap.columns.get("ean")
    if ean_col is None:
        raise ValueError(f"{path}: no EAN column found via header synonyms")

    rows = []
    for row in ws.iter_rows(min_row=tmap.header_row + 1):
        if row[ean_col - 1].value in (None, ""):
            continue
        record: dict[str, object] = {"_row": row[0].row}
        for canonical, col in tmap.columns.items():
            record[canonical] = row[col - 1].value
        rows.append(record)
    return rows
