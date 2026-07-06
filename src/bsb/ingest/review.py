"""Ingest a Felina-reviewed sheet and grow the brand color-code lexicon
(Oli 2026-07-06).

We hand Felina the emitted review copy: color_code auto-PROPOSALS are yellow,
disagreements/no-signal are red. She confirms, corrects, or fills. On return
this reads her final color_code (Data sheet) against our proposal (Provenance
sheet, unchanged by her) and:

- proposal confirmed unchanged  -> her/our value enters the brand lexicon
  (Felina-confirmed), counted "accepted" for that proposal source;
- proposal corrected            -> HER value enters the lexicon; the change is
  logged as a correction against the proposal source;
- red cell she fills            -> a normal Felina-decided lexicon entry.

Confirmed entries are appended to config/lexicons/{brand}.yaml (keyed on the
canonical color_name), which load_brands merges into shade_lexicon — so the
NEXT order maps deterministically. brands.yaml is never rewritten; a shade
already in the lexicon is left untouched (idempotent). The per-source
correction rate is the auto-proposer's quality metric: a high-correction
source gets demoted or dropped on the data.
"""

from collections import defaultdict
from pathlib import Path

import openpyxl
import yaml
from pydantic import BaseModel, Field

# proposal-source tags parsed from the color_code provenance note/rule
_SOURCES = ("two_signal", "hex_only", "word_only", "rule", "none")


def _classify_source(notes: str) -> str:
    n = (notes or "").lower()
    # reds first: a disagreement note shows BOTH signals (mentions swatch_hex),
    # so it must be matched before the hex-only proposal check below.
    if "signals disagree" in n or "needs human decision" in n:
        return "none"  # withheld disagreement — no proposal
    if "no color-code rule matched" in n or "fail closed" in n:
        return "none"  # red, no signal
    if "two signals agree" in n or "two_signals_agree" in n:
        return "two_signal"
    if "word_over_low_chroma_hex" in n or "word signal used" in n:
        return "word_only"
    if "proposed from swatch hex" in n or "swatch hex low-chroma" in n:
        return "hex_only"
    if "proposed from shade name" in n or "color_word" in n:
        return "word_only"
    return "rule"  # foundation/clear/etc — a confirmed rule, not a proposal


class ReviewOutcome(BaseModel):
    accepted: dict[str, int] = Field(default_factory=lambda: defaultdict(int))
    corrected: dict[str, int] = Field(default_factory=lambda: defaultdict(int))
    felina_filled_reds: int = 0
    new_entries: list[dict] = Field(default_factory=list)
    skipped_existing: int = 0

    def correction_report(self) -> list[str]:
        lines = []
        for src in ("two_signal", "hex_only", "word_only"):
            a, c = self.accepted.get(src, 0), self.corrected.get(src, 0)
            total = a + c
            if total:
                rate = c / total
                lines.append(f"  {src:11}: {a} accepted, {c} corrected → {rate:.0%} correction")
        return lines


def _read_sheets(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # Data sheet: her final values, keyed by EAN
    ws = wb.worksheets[0]
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(header)}
    ean_i = next((col[h] for h in ("EAN Code", "EAN", "ean") if h in col), 0)
    cc_i = col.get("Boozt Color code")
    cn_i = col.get("Color Name")
    final: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[ean_i] in (None, ""):
            continue
        final[str(row[ean_i])] = {
            "color_code": row[cc_i] if cc_i is not None else None,
            "color_name": row[cn_i] if cn_i is not None else None,
        }
    # Provenance: our proposal, unchanged by her
    prop: dict[str, dict] = {}
    if "Provenance" in wb.sheetnames:
        prov = wb["Provenance"]
        rows = prov.iter_rows(values_only=True)
        ph = {str(h): i for i, h in enumerate(next(rows))}
        for r in rows:
            if r[ph["field"]] == "color_code":
                prop[str(r[ph["ean"]])] = {"value": r[ph["value"]], "notes": r[ph["notes"]]}
    wb.close()
    return final, prop


def ingest_review(
    reviewed_path: str | Path,
    brand_key: str,
    reviewer: str = "Felina",
    date: str = "",
    config_dir: Path | None = None,
) -> ReviewOutcome:
    from bsb.config import DEFAULT_CONFIG_DIR, load_brands

    config_dir = config_dir or DEFAULT_CONFIG_DIR
    final, prop = _read_sheets(Path(reviewed_path))
    out = ReviewOutcome()

    lex_path = config_dir / "lexicons" / f"{brand_key}.yaml"
    # entries we rewrite live in the lexicon FILE; the skip-set is the brand's
    # FULL current lexicon (curated brands.yaml + file) so an already-confirmed
    # shade — curated or previously ingested — is never shadowed (idempotent).
    existing_entries = []
    if lex_path.exists():
        existing_entries = (yaml.safe_load(lex_path.read_text()) or {}).get("entries") or []
    merged_lexicon = (load_brands(config_dir).get(brand_key) or {}).get("shade_lexicon") or []
    existing_keys = {str(e.get("shade", "")).casefold() for e in merged_lexicon}

    def _norm(v):
        return str(v).strip() if v not in (None, "") else None

    for ean, fin in final.items():
        her = _norm(fin.get("color_code"))
        shade = _norm(fin.get("color_name"))
        if not her or not shade:
            continue  # she left the code blank, or no shade to key on
        key = shade.casefold()
        our = prop.get(ean, {})
        source = _classify_source(str(our.get("notes") or ""))
        our_val = _norm(our.get("value"))

        if source in ("two_signal", "hex_only", "word_only"):
            if our_val == her:
                out.accepted[source] += 1
                status = "accepted"
            else:
                out.corrected[source] += 1
                status = f"corrected_from_{source}:{our_val}"
        elif source == "none":
            out.felina_filled_reds += 1
            status = "felina_decided"
        else:  # rule-confirmed cell; only record if she changed it
            if our_val == her:
                continue
            status = f"corrected_rule:{our_val}"

        if key in existing_keys:
            out.skipped_existing += 1
            continue
        existing_keys.add(key)
        out.new_entries.append(
            {
                "shade": key,
                "code": int(her) if str(her).isdigit() else her,
                "decided_by": reviewer,
                "date": date,
                "source": status,
            }
        )

    _write_lexicon(lex_path, existing_entries, out.new_entries)
    return out


def _write_lexicon(lex_path: Path, existing: list[dict], new: list[dict]) -> None:
    if not new:
        return
    lex_path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "version": 1,
        "note": "Felina-confirmed shade->color_code (grown by `bsb ingest-review`); "
        "keyed on canonical color_name. Merged into shade_lexicon by load_brands.",
        "entries": existing + new,
    }
    lex_path.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True))
