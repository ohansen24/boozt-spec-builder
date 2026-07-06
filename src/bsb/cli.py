"""bsb CLI (build kit section 7): run, resolve, report.

Phase 0 wires the pure-local pipeline: ingest -> categorize -> normalize ->
emit. Fields that need web sources carry status NOT_FOUND until the resolve /
extract stages exist (Phase 1).
"""

import time
from datetime import UTC, datetime
from pathlib import Path

import click

from bsb.config import DEFAULT_CONFIG_DIR, load_brands, load_header_synonyms, load_rules
from bsb.emit.writer import RunSummary, write_output
from bsb.ingest.odm import parse_odm
from bsb.pipeline import build_records


class _Progress:
    """Periodic progress line for long runs: every `every` items prints
    `label n/total | Xm Ys | cache C | fc <credits>`. Per-item messages still
    flow to the log; the summary is the watch-friendly heartbeat."""

    def __init__(self, total, label, *, fetcher=None, firecrawl=None, every=10):
        self.total = total
        self.label = label
        self.every = every
        self.fetcher = fetcher
        self.firecrawl = firecrawl
        self.n = 0
        self.start = time.monotonic()

    def __call__(self, msg=None):
        if msg:
            click.echo("  " + msg)
        self.n += 1
        if self.n % self.every and self.n != self.total:
            return
        el = int(time.monotonic() - self.start)
        bits = [f"{self.label} {self.n}/{self.total}", f"{el // 60}m{el % 60:02d}s"]
        if self.fetcher is not None:
            bits.append(f"cache {self.fetcher.stats['cache_hits']}")
        if self.firecrawl is not None:
            u = self.firecrawl.usage
            bits.append(f"fc {u['scrapes']}scr/{u['searches']}srch")
        click.echo("  >> " + " | ".join(bits))


@click.group()
def main() -> None:
    """Boozt Spec Builder: ODM + blank Boozt template -> filled, validated
    product data sheet with per-field provenance."""


DEFAULT_TEMPLATE = Path(__file__).resolve().parents[2] / "data/templates/boozt_beauty_master.xlsx"


@main.command()
@click.option("--odm", "odm_path", required=True, type=click.Path(exists=True, dir_okay=False))
@click.option(
    "--template",
    "template_path",
    type=click.Path(exists=True, dir_okay=False),
    default=str(DEFAULT_TEMPLATE),
    show_default=True,
    help="Boozt template copy target (explicit path wins over the master template)",
)
@click.option(
    "--brand",
    "brand_key",
    default=None,
    help="Brand key (default: auto-detect from the ODM order number)",
)
@click.option("--out", "out_path", required=True, type=click.Path(dir_okay=False))
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
@click.option(
    "--cache-dir",
    type=click.Path(file_okay=False, path_type=Path),
    default=Path("cache"),
    show_default=True,
)
@click.option("--resolve/--no-resolve", "do_resolve", default=False, help="Fetch brand-site data")
@click.option(
    "--validate/--no-validate", "do_validate", default=True, help="Validator pass (with --resolve)"
)
@click.option("--bases", "bases_csv", default=None, help="Comma-separated base-name filter (pilot)")
@click.option("--show-variants", is_flag=True, help="Print per-master variant tables")
@click.option("--sample-provenance", default=0, type=int, help="Print N random provenance rows")
@click.option(
    "--allow-size-anomalies",
    is_flag=True,
    help="Do not stop on site-vs-ODM size mismatches; ship yellow with notes (kit 6.5)",
)
def run(
    odm_path: str,
    template_path: str,
    brand_key: str,
    out_path: str,
    config_dir: Path,
    cache_dir: Path,
    do_resolve: bool,
    do_validate: bool,
    bases_csv: str | None,
    show_variants: bool,
    sample_provenance: int,
    allow_size_anomalies: bool,
) -> None:
    """Fill a Boozt template from an ODM (--resolve adds brand-site data and
    the validator pass; anomalies stop the run before emit)."""
    brands = load_brands(config_dir)
    rules = load_rules(config_dir)
    synonyms = load_header_synonyms(config_dir)

    odm = _ingest_order(odm_path, synonyms)

    if brand_key is None:
        from bsb.config import brand_for_order

        brand_key = brand_for_order(odm.order_number, brands)
        if brand_key is None:
            raise click.BadParameter(
                f"cannot auto-detect brand from order {odm.order_number!r} — pass --brand"
            )
        click.echo(f"Brand auto-detected from order {odm.order_number}: {brand_key}")
    brand_key = brand_key.lower()
    if brand_key not in brands:
        raise click.BadParameter(f"unknown brand {brand_key!r}; known: {', '.join(sorted(brands))}")
    brand_cfg = brands[brand_key]
    if brand_cfg.get("out_of_scope"):
        raise click.ClickException(
            f"brand {brand_key!r} is out of scope (different Boozt template/guide)"
        )

    bases_filter = None
    if bases_csv:
        bases_filter = {b.strip() for b in bases_csv.split(",") if b.strip()}
        odm.rows = [r for r in odm.rows if r.base_name in bases_filter]
        click.echo(f"Base filter: {len(odm.rows)} rows across {len(bases_filter)} bases")

    records = build_records(odm, brand_key, brands, rules, odm_path)

    run_meta = {
        "run at": datetime.now(UTC).isoformat(timespec="seconds"),
        "odm": odm_path,
        "template": template_path,
        "brand": brand_key,
        "guide version": rules["guide_version"],
        "ean length profile": ", ".join(
            f"{length}-digit: {count}" for length, count in sorted(odm.length_profile.items())
        ),
        "_ingest_issues": odm.issues,
    }

    if not do_resolve:
        run_meta["phase"] = "0 (pure local — no web sources; web-dependent fields are NOT_FOUND)"
        summary = write_output(template_path, out_path, records, synonyms, run_meta)
        click.echo("EAN length profile: " + run_meta["ean length profile"])
        _print_summary(summary)
        return

    _run_resolved(
        odm,
        records,
        brand_key,
        brand_cfg,
        rules,
        synonyms,
        template_path,
        out_path,
        cache_dir,
        config_dir,
        run_meta,
        do_validate,
        show_variants,
        sample_provenance,
        allow_size_anomalies,
    )


def _run_resolved(
    odm,
    records,
    brand_key,
    brand_cfg,
    rules,
    synonyms,
    template_path,
    out_path,
    cache_dir,
    config_dir,
    run_meta,
    do_validate,
    show_variants,
    sample_provenance,
    allow_size_anomalies,
) -> None:
    import random

    from bsb.fetch.cache import EanCache, HttpCache
    from bsb.fetch.ladder import PlaywrightSession, PoliteFetcher
    from bsb.pipeline import apply_resolution
    from bsb.resolve.adapters.sfcc import SfccAdapter
    from bsb.resolve.orchestrator import resolve_order
    from bsb.resolve.validators import IncidecoderWeak, LookfantasticValidator, cache_lf_hit

    adapter_type = brand_cfg.get("adapter")
    if adapter_type not in ("nars_sfcc", "sfcc", "shopify"):
        raise click.ClickException(
            f"--resolve: no resolve-capable adapter configured for brand {brand_key!r} "
            f"(adapter={adapter_type!r})"
        )

    http_cache = HttpCache(cache_dir)
    ean_cache = EanCache(cache_dir)
    fetcher = PoliteFetcher(http_cache)
    playwright = PlaywrightSession(http_cache, fetcher.limiter)

    try:
        if adapter_type == "shopify":
            from bsb.resolve.adapters.shopify import ShopifyAdapter
            from bsb.resolve.orchestrator import resolve_order_shopify

            adapter = ShopifyAdapter(fetcher, brand_cfg, ean_cache)
            click.echo(f"Resolving {len(odm.rows)} EANs (Shopify barcode anchor)…")
            resolution = resolve_order_shopify(
                odm,
                adapter,
                brand_cfg,
                progress=_Progress(len(odm.rows), "resolve", fetcher=fetcher),
            )
        else:
            adapter = SfccAdapter(fetcher, brand_cfg, ean_cache, playwright)
            click.echo(f"Resolving {len(odm.rows)} EANs master-first…")
            resolution = resolve_order(
                odm, adapter, progress=_Progress(len(odm.rows), "resolve", fetcher=fetcher)
            )
        counts = resolution.counts()
        click.echo(
            f"Resolve done: {counts['resolved_ok']}/{counts['eans']} ok, "
            f"{counts['masters_found']} masters"
        )

        if resolution.blocking_anomalies:
            click.echo("\n!! BLOCKING ANOMALIES — stopping before validators/emit:")
            for a in resolution.anchor_rejections:
                click.echo(f"  anchor rejection: {a}")
            for a in resolution.missing_shades:
                click.echo(f"  missing shade:    {a}")
            raise SystemExit(2)
        for failure in resolution.master_failures:
            click.echo(f"  (non-blocking) master not found: {failure}")
        for warning in resolution.swatch_warnings:
            click.echo(f"  (warning) {warning}")
        if resolution.swatch_warnings:
            run_meta["swatch warnings"] = " | ".join(resolution.swatch_warnings)

        # key validators by the RESOLVED master id, not base_name — spec-sheet
        # (Shopify) orders have no ODM name to group on, so base_name is empty
        lf_by_master: dict = {}
        inci_by_master: dict = {}
        rep_ean_by_master: dict = {}
        for ean, res in resolution.by_ean.items():
            if res.ok and res.master is not None:
                rep_ean_by_master.setdefault(res.master.master_id, (ean, res.master))
        generic_by_line: dict = {}  # product_name -> (LfProduct-shaped, (inci_text, url))
        if do_validate:
            lf = LookfantasticValidator(fetcher, playwright)
            inci_weak = IncidecoderWeak(fetcher)
            click.echo(f"\nValidator pass over {len(rep_ean_by_master)} masters…")
            lf_tick = _Progress(len(rep_ean_by_master), "LF-validate", fetcher=fetcher)
            for master_id, (first_ean, master) in rep_ean_by_master.items():
                product = lf.find_product(first_ean)
                lf_by_master[master_id] = product
                inci_by_master[master_id] = inci_weak.find_inci(
                    str(brand_cfg.get("display_name", brand_key)), master.product_name
                )
                hit = f"{len(product.by_barcode)} barcodes @ {product.url}" if product else "no hit"
                lf_tick(f"LF {master.product_name[:40]}: {hit}")

            # validator-of-last-resort: if the configured pool found nothing for
            # this brand, let the generic resolver discover retailer families
            # (search -> GTIN-anchored PDP). Deduped per product line (INCI and
            # name are shared across a line's sizes); cache-first, capped.
            if not any(lf_by_master.values()):
                generic_by_line = _generic_validator_pass(
                    resolution,
                    brand_cfg,
                    brand_key,
                    fetcher,
                    config_dir,
                    run_meta,
                    progress=lambda m: click.echo("  " + m),
                )

        conflict_cells = 0
        compared_cells = 0
        size_anomalies: list[str] = []
        by_ean_row = {r.ean12: r for r in odm.rows}
        for record in records:
            row = by_ean_row[record.ean12]
            resolved = resolution.by_ean.get(record.ean12)
            master_id = resolved.master.master_id if (resolved and resolved.master) else None
            lf_product = lf_by_master.get(master_id)
            weak = inci_by_master.get(master_id)
            retailer_inci = None
            if lf_product is None and resolved and resolved.master:
                gen = generic_by_line.get(resolved.master.product_name)
                if gen:
                    lf_product, retailer_inci = gen
            size_anomalies += apply_resolution(
                record, row, resolved, brand_cfg, rules, lf_product, weak, retailer_inci
            )
            for field in ("style_name", "color_name", "size"):
                fv = getattr(record, field)
                if fv.secondary is not None or fv.status == "CONFLICT":
                    compared_cells += 1
                    if fv.status == "CONFLICT":
                        conflict_cells += 1

            if lf_product is not None and row.ean12 in lf_product.by_barcode:
                cache_lf_hit(ean_cache, row.gtin13, lf_product, row.ean12)

        if size_anomalies:
            click.echo(
                "\n!! SIZE ANOMALIES vs ODM hints"
                + (
                    " (allowed; shipping yellow):"
                    if allow_size_anomalies
                    else " — stopping before emit:"
                )
            )
            for a in size_anomalies:
                click.echo(f"  {a}")
            if not allow_size_anomalies:
                raise SystemExit(2)
            run_meta["size anomalies (accepted)"] = " | ".join(size_anomalies)

        from bsb.config import load_order_overrides
        from bsb.pipeline import apply_order_overrides

        if odm.order_number:
            override_path = f"config/order_overrides/{odm.order_number}.yaml"
            entries = load_order_overrides(odm.order_number)
            if entries:
                applied = apply_order_overrides(records, entries, override_path)
                click.echo(f"\nOrder overrides ({override_path}): {applied} cells replaced")
                run_meta["order overrides"] = f"{applied} cells from {override_path}"
                # overridden size cells are settled decisions, not anomalies
                overridden = {(str(ean), e["field"]) for e in entries for ean in e.get("eans", [])}
                size_anomalies = [
                    a for a in size_anomalies if (a.split(" ")[0], "size") not in overridden
                ]

        conflict_rate = (conflict_cells / compared_cells) if compared_cells else 0.0
        click.echo(
            f"\nValidator comparisons: {compared_cells} cells, {conflict_cells} conflicts "
            f"({conflict_rate:.1%})"
        )
        if conflict_rate > 0.10:
            click.echo("!! validator conflict rate above 10% — stopping before emit")
            raise SystemExit(2)

        if show_variants:
            _print_variant_tables(resolution, lf_by_master, odm)

        shade_fmt = brand_cfg.get("shade_format") or {}
        if shade_fmt.get("pending_note"):
            run_meta["shade format"] = shade_fmt["pending_note"]
        run_meta["phase"] = "1 (brand adapter + validator pass)"
        run_meta["validators"] = (
            "lookfantastic (GTIN-anchored), incidecoder (weak, notes only); "
            "douglas/boots/flaconi/sephora disabled: bot-walled at httpx and Playwright rungs"
        )
        run_meta["inci casing"] = "as published by brand (ALL CAPS); separators normalized to comma"

        summary = write_output(template_path, out_path, records, synonyms, run_meta)
        _print_summary(summary)
        _print_queue_by_reason(summary)
        if sample_provenance:
            _print_provenance_sample(records, sample_provenance, random.Random(20260702))
    finally:
        fetcher.close()
        playwright.close()


def _generic_validator_pass(
    resolution, brand_cfg, brand_key, fetcher, config_dir, run_meta, progress
):
    """Generic resolver as validator-of-last-resort (pool had zero coverage).
    Per distinct product line: search -> GTIN-anchored retailer PDP -> a
    validator family (name/shade/size, as an LfProduct) + retailer INCI.
    Records working families into validators.yaml; reports Firecrawl credits."""
    from bsb.fetch.firecrawl import FirecrawlClient
    from bsb.resolve.generic import GenericResolver
    from bsb.resolve.validators import LfProduct, LfVariant

    firecrawl = FirecrawlClient(fetcher.cache, fetcher.limiter)
    if not firecrawl.available:
        progress("generic validator unavailable (no FIRECRAWL_API_KEY) — skipping")
        return {}
    resolver = GenericResolver(fetcher, firecrawl)
    brand = str(brand_cfg.get("search_brand") or brand_cfg.get("display_name") or brand_key)

    # one representative EAN per product line (INCI/name shared across sizes)
    rep_by_line: dict[str, str] = {}
    for _ean, res in resolution.by_ean.items():
        if res.ok and res.master is not None:
            rep_by_line.setdefault(res.master.product_name, res.ean12)

    usage0 = firecrawl.snapshot_usage()
    generic_by_line: dict = {}
    families: dict[str, int] = {}
    progress(f"validator-of-last-resort: generic resolver over {len(rep_by_line)} product lines…")
    tick = _Progress(len(rep_by_line), "generic-val", fetcher=fetcher, firecrawl=firecrawl)
    for line, ean12 in rep_by_line.items():
        gtin13 = ean12 if len(ean12) == 13 else "0" + ean12
        hits = resolver.resolve(gtin13, ean12, brand, max_pages=3)
        anchored = [h for h in hits if h.gtin_anchored]
        for h in anchored:
            families[h.family] = families.get(h.family, 0) + 1
        best_named = next((h for h in anchored if h.name), None)
        best_inci = next((h for h in anchored if h.inci), None)
        lf_shaped = None
        if best_named:
            lf_shaped = LfProduct(
                url=best_named.url,
                product_name=best_named.name,
                size_text=best_named.size,
                by_barcode={ean12: LfVariant(barcode=ean12, shade=best_named.color)},
            )
        retailer_inci = (best_inci.inci, best_inci.url) if best_inci else None
        if lf_shaped or retailer_inci:
            generic_by_line[line] = (lf_shaped, retailer_inci)
        tick(
            f"{line[:38]}: "
            + (f"{len(anchored)} anchored" if anchored else "no anchored family")
            + (" +INCI" if retailer_inci else "")
        )

    used = firecrawl.usage_since(usage0)
    run_meta["generic validator credits"] = (
        f"{used['scrapes']} scrapes, {used['searches']} searches "
        f"({used['search_results']} results), {used['cache_hits']} cache hits"
    )
    # record working families as this brand's validator pool for next time
    if families:
        top = sorted(families, key=lambda f: -families[f])
        run_meta["discovered validator families"] = ", ".join(f"{f}({families[f]})" for f in top)
        _record_validator_pool(config_dir, brand_key, top, progress)
    return generic_by_line


def _record_validator_pool(config_dir, brand_key, families, progress) -> None:
    """Append the discovered retailer families to config/validators.yaml as
    the brand's pool, so the next order tries them directly."""
    import yaml

    path = config_dir / "validators.yaml"
    try:
        data = yaml.safe_load(path.read_text()) or {}
    except FileNotFoundError:
        data = {}
    discovered = data.setdefault("discovered_pools", {})
    if discovered.get(brand_key) != families:
        discovered[brand_key] = families
        path.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True))
        progress(f"recorded validator pool for {brand_key}: {', '.join(families[:5])}")


def _print_variant_tables(resolution, lf_by_master, odm) -> None:
    click.echo("\n=== per-master variant tables ===")
    for base, master in resolution.masters.items():
        lf = lf_by_master.get(master.master_id)
        click.echo(
            f"\nmaster {master.master_id}  ({master.product_name})  "
            f"{len(master.shade_by_gtin)} shades on site"
        )
        click.echo(f"  pdp  {master.pdp_url}")
        if master.inci_text:
            click.echo(
                f"  inci {master.inci_text[:100]}… (shade {master.inci_selected_gtin} selected)"
            )
        if lf:
            click.echo(f"  LF   {lf.url} ({len(lf.by_barcode)} barcodes)")
        for row in [r for r in odm.rows if r.base_name == base]:
            entry = resolution.by_ean.get(row.ean12)
            variant = entry.variant if entry else None
            lf_variant = lf.by_barcode.get(row.ean12) if lf else None
            status = "ok" if entry and entry.ok else f"FAIL: {entry.error if entry else '?'}"
            shade = variant.shade if variant else "-"
            size = variant.size_text if variant else "-"
            lf_note = f" | LF: {lf_variant.shade}" if lf_variant else " | LF: -"
            click.echo(f"    {row.ean12}  {status:4} shade={shade!r} size={size!r}{lf_note}")


def _print_queue_by_reason(s: RunSummary) -> None:
    def reason_of(item) -> str:
        note = (item.notes or "").split(";")[0].strip()
        return f"{item.field}: {note[:70] if note else item.status}"

    click.echo("\nReview queue by reason:")
    groups: dict[str, int] = {}
    for item in s.review_red + s.review_yellow:
        key = ("RED " if item in s.review_red else "YEL ") + reason_of(item)
        groups[key] = groups.get(key, 0) + 1
    for key, count in sorted(groups.items(), key=lambda kv: (-kv[1], kv[0]))[:20]:
        click.echo(f"  {count:4d}  {key}")


def _print_provenance_sample(records, n: int, rng) -> None:
    click.echo(f"\n{n} random provenance rows (browser spot-checks):")
    pool = []
    for record in records:
        from bsb.models import ProductRecord

        for field in ProductRecord.field_values():
            fv = getattr(record, field)
            if fv.value and fv.primary:
                pool.append((record.ean12, field, fv))
    for ean, field, fv in rng.sample(pool, min(n, len(pool))):
        click.echo(f"  {ean} {field} = {fv.value!r} [{fv.status}]")
        click.echo(f"      primary   {fv.primary.url}")
        if fv.secondary:
            click.echo(f"      secondary {fv.secondary.url}")
        click.echo(f"      snippet   {fv.primary.snippet[:100]}")


def _ingest_order(odm_path: str, synonyms: dict):
    """Dispatch on input shape: a Buying Labs ODM (Barcode/Name/QTY header
    block) vs a Boozt-template "Specs" sheet (EAN Code header in row 1)."""
    from bsb.ingest.odm import parse_spec_sheet

    try:
        return parse_odm(odm_path)
    except ValueError:
        return parse_spec_sheet(odm_path, synonyms)


def _print_summary(s: RunSummary) -> None:
    click.echo(f"\nWrote {s.out_path} ({s.records} rows)")
    if s.cleared_template_rows:
        click.echo(
            f"  ! template contained {s.cleared_template_rows} pre-existing data rows — "
            "cleared in the output copy"
        )
    for issue in s.ingest_issues:
        click.echo(f"  ! ingest: {issue}")

    click.echo("\nStatus totals (written cells):")
    for status, count in sorted(s.status_totals.items()):
        click.echo(f"  {status:14} {count}")

    click.echo("\nCategories:")
    for category, count in sorted(s.category_totals.items(), key=lambda kv: -kv[1]):
        click.echo(f"  {category:20} {count}")

    if s.verify_at_receipt:
        click.echo(
            f"\nVERIFY AT RECEIPT ({len(s.verify_at_receipt)} cells — warehouse checkpoint):"
        )
        for item in s.verify_at_receipt:
            click.echo(f"  {item.ean} {item.field} = {item.value!r}")
    click.echo(
        f"\nReview queue: {len(s.review_red)} red, {len(s.review_yellow)} yellow "
        "(see 'Run report' sheet)"
    )
    by_field: dict[str, int] = {}
    for item in s.review_red:
        by_field[item.field] = by_field.get(item.field, 0) + 1
    if by_field:
        click.echo("  red by field: " + ", ".join(f"{f}={n}" for f, n in sorted(by_field.items())))
    yellow_by_field: dict[str, int] = {}
    for item in s.review_yellow:
        yellow_by_field[item.field] = yellow_by_field.get(item.field, 0) + 1
    if yellow_by_field:
        click.echo(
            "  yellow by field: "
            + ", ".join(f"{f}={n}" for f, n in sorted(yellow_by_field.items()))
        )
    if s.unknown_headers:
        click.echo("\nUnknown template headers (left untouched): " + "; ".join(s.unknown_headers))
    if s.missing_fields:
        click.echo(
            "Canonical fields absent from template (skipped): " + "; ".join(s.missing_fields)
        )


ANSWER_KEY_FIXTURES = {
    "aesop": "tests/fixtures/aesop_final.xlsx",
    "olaplex": "tests/fixtures/olaplex_final.xlsx",
    "svr": "data/inbox/blank_template.xlsx",  # Felina's finished SVR rows
}


@main.command("probe-brand")
@click.argument("brand_key")
@click.option("--eans", "eans_csv", default=None, help="Comma-separated sample EANs")
@click.option("--samples", default=5, type=int, show_default=True)
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
@click.option(
    "--cache-dir",
    type=click.Path(file_okay=False, path_type=Path),
    default=Path("cache"),
    show_default=True,
)
def probe_brand_cmd(
    brand_key: str, eans_csv: str | None, samples: int, config_dir: Path, cache_dir: Path
) -> None:
    """Probe a brand before building: platform, EAN addressability, on-site
    INCI, fixtures + a draft brands.yaml entry. No gates, no bespoke code."""

    from bsb.fetch.cache import HttpCache
    from bsb.fetch.ladder import PoliteFetcher
    from bsb.probe import probe_brand

    brands = load_brands(config_dir)
    brand_key = brand_key.lower()
    if brand_key not in brands:
        raise click.BadParameter(f"unknown brand {brand_key!r}")

    sample_eans: list[str] = []
    if eans_csv:
        sample_eans = [e.strip() for e in eans_csv.split(",") if e.strip()]
    elif brand_key in ANSWER_KEY_FIXTURES:
        from openpyxl import load_workbook

        ws = load_workbook(ANSWER_KEY_FIXTURES[brand_key])["Data sheet"]
        sample_eans = [
            str(ws.cell(row=r, column=1).value)
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ][:samples]

    fetcher = PoliteFetcher(HttpCache(cache_dir))
    try:
        report = probe_brand(
            brand_key, brands[brand_key], fetcher, sample_eans, Path("tests/fixtures/probes")
        )
    finally:
        fetcher.close()

    out = Path(f"data/out/probe_{brand_key}.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(report.model_dump_json(indent=1))
    click.echo(f"PROBE {brand_key}: platform={report.platform} domain={report.domain}")
    click.echo(f"  ean_addressable: {report.ean_addressable}  ({report.ean_evidence})")
    if report.barcodes_in_catalog:
        click.echo(f"  catalog barcodes: {report.barcodes_in_catalog}")
    click.echo(f"  inci_on_site: {report.inci_on_site}  ({report.inci_evidence})")
    for note in report.notes:
        click.echo(f"  note: {note}")
    click.echo(f"  fixtures: {len(report.fixtures)} | report: {out}")
    click.echo("  draft brands.yaml:")
    for line in report.draft_yaml().splitlines():
        click.echo(f"    {line}")


@main.command("golden")
@click.option("--brand", "brand_key", required=True)
@click.option("--limit", default=None, type=int, help="Compare only the first N rows")
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
@click.option(
    "--cache-dir",
    type=click.Path(file_okay=False, path_type=Path),
    default=Path("cache"),
    show_default=True,
)
def golden_cmd(brand_key: str, limit: int | None, config_dir: Path, cache_dir: Path) -> None:
    """Golden comparison: resolve offline from probe cache and diff
    auto-fillable fields against Felina's finished sheet for this brand."""
    from bsb.golden import golden_compare

    brand_key = brand_key.lower()
    if brand_key not in ANSWER_KEY_FIXTURES:
        raise click.BadParameter(
            f"no answer key for {brand_key!r}; have: {', '.join(ANSWER_KEY_FIXTURES)}"
        )
    brands = load_brands(config_dir)
    synonyms = load_header_synonyms(config_dir)
    result = golden_compare(
        brand_key,
        brands[brand_key],
        ANSWER_KEY_FIXTURES[brand_key],
        synonyms,
        cache_dir,
        limit=limit,
    )
    click.echo(f"GOLDEN {brand_key}: {result.resolved}/{result.rows} rows resolved")
    for note in result.notes:
        click.echo(f"  note: {note}")
    for field, stats in result.fields.items():
        if stats.comparable or stats.tool_missing or stats.theirs_missing:
            click.echo(
                f"  {field:12} agreement {stats.rate:6.1%} "
                f"(agree {stats.agree} + format-only {stats.format_only} "
                f"/ {stats.comparable} comparable; tool-missing {stats.tool_missing}, "
                f"felina-missing {stats.theirs_missing})"
            )
    if result.disagreements:
        click.echo(f"  disagreements ({len(result.disagreements)}):")
        for d in result.disagreements[:20]:
            click.echo(f"    {d}")
    out = Path(f"data/out/golden_{brand_key}.json")
    out.write_text(result.model_dump_json(indent=1))
    click.echo(f"  report: {out}")


@main.command("ingest-portal-errors")
@click.option("--sheet", "sheet_path", required=True, type=click.Path(exists=True, dir_okay=False))
@click.option("--order", "order_id", required=True)
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
def ingest_portal_errors(sheet_path: str, order_id: str, config_dir: Path) -> None:
    """Turn Boozt portal rejections (the 'Boozt Errors' column) into DRAFT
    order overrides for review — the portal is the final QA stage and its
    feedback compounds like everything else. Nothing is applied without a
    human filling in the corrected values."""
    from bsb.portal import collect_portal_errors, draft_overrides

    synonyms = load_header_synonyms(config_dir)
    errors = collect_portal_errors(sheet_path, synonyms)
    if not errors:
        click.echo("No portal errors found — nothing to draft.")
        return
    click.echo(f"{len(errors)} portal errors:")
    for e in errors:
        click.echo(f"  {e.ean} [{e.field_guess}] {e.error[:90]}")
    path = draft_overrides(order_id, errors, config_dir / "order_overrides")
    click.echo(f"\nDraft written: {path}")
    click.echo(
        "Review it, fill the FIXME values, merge accepted entries into "
        f"config/order_overrides/{order_id}.yaml — recurring patterns belong "
        "in rules/lexicon config instead."
    )


@main.command("compare-external")
@click.option(
    "--theirs", "theirs_path", required=True, type=click.Path(exists=True, dir_okay=False)
)
@click.option("--order", "order_id", required=True)
@click.option("--ours", "ours_path", type=click.Path(exists=True, dir_okay=False), default=None)
@click.option("--out", "out_path", type=click.Path(dir_okay=False), default=None)
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
def compare_external(
    theirs_path: str,
    order_id: str,
    ours_path: str | None,
    out_path: str | None,
    config_dir: Path,
) -> None:
    """Diff an externally hand-filled sheet against the tool's output and
    emit the review-session workbook. Nothing is ingested here: every
    DISAGREE row needs a human Decision first."""
    from bsb.compare import compare_sheets, write_comparison

    ours_path = ours_path or f"data/out/{order_id}_review.xlsx"
    out_path = out_path or f"data/out/comparison_{order_id}.xlsx"
    synonyms = load_header_synonyms(config_dir)

    result = compare_sheets(ours_path, theirs_path, synonyms)
    write_comparison(result, out_path, ours_path, theirs_path)

    pct = (result.agree / result.compared_cells * 100) if result.compared_cells else 0.0
    click.echo(f"AGREEMENT: {result.agree} of {result.compared_cells} cells identical ({pct:.1f}%)")
    counts: dict[str, int] = {}
    for d in result.differences:
        counts[d.classification] = counts.get(d.classification, 0) + 1
    click.echo(
        f"  DISAGREE {counts.get('DISAGREE', 0)} | FORMAT_ONLY {counts.get('FORMAT_ONLY', 0)} "
        f"| FELINA_ONLY {counts.get('FELINA_ONLY', 0)} | TOOL_ONLY {counts.get('TOOL_ONLY', 0)} "
        f"| whitespace-dirty cells in her file (normalized away): "
        f"{result.theirs_whitespace_dirty}"
    )
    click.echo(f"Wrote {out_path}")

    format_only = [d for d in result.differences if d.classification == "FORMAT_ONLY"]
    if format_only:
        click.echo(f"\nFORMAT_ONLY — surfaced, no decision required ({len(format_only)}):")
        for d in format_only:
            click.echo(f"  {d.ean} {d.field}: tool {d.ours!r} vs felina {d.theirs!r} [{d.note}]")

    disagreements = result.disagreements
    if disagreements:
        click.echo(f"\nDISAGREE — session agenda ({len(disagreements)}):")
        for d in disagreements:
            note = f"   [{d.note}]" if d.note else ""
            click.echo(f"  {d.ean} {d.field}: tool {d.ours!r} vs felina {d.theirs!r}{note}")
            if d.provenance_url:
                click.echo(f"      tool source: {d.provenance_url}")


@main.command()
@click.option("--gtin", required=True, help="GTIN-13 (0 + ODM ean12)")
@click.option("--brand", "brand_key", required=True)
@click.option("-v", "--verbose", is_flag=True)
@click.option(
    "--config",
    "config_dir",
    type=click.Path(exists=True, file_okay=False, path_type=Path),
    default=DEFAULT_CONFIG_DIR,
    show_default=True,
)
@click.option(
    "--cache-dir",
    type=click.Path(file_okay=False, path_type=Path),
    default=Path("cache"),
    show_default=True,
)
def resolve(gtin: str, brand_key: str, verbose: bool, config_dir: Path, cache_dir: Path) -> None:
    """Single-item debug resolve: master discovery + GTIN-anchored variation."""
    from bsb.fetch.cache import EanCache, HttpCache
    from bsb.fetch.ladder import PlaywrightSession, PoliteFetcher
    from bsb.resolve.adapters.sfcc import SfccAdapter

    brands = load_brands(config_dir)
    brand_key = brand_key.lower()
    brand_cfg = brands.get(brand_key)
    if not brand_cfg or brand_cfg.get("adapter") not in ("nars_sfcc", "sfcc"):
        raise click.BadParameter(f"no adapter configured for brand {brand_key!r}")

    http_cache = HttpCache(cache_dir)
    fetcher = PoliteFetcher(http_cache)
    playwright = PlaywrightSession(http_cache, fetcher.limiter)
    adapter = SfccAdapter(fetcher, brand_cfg, EanCache(cache_dir), playwright)

    try:
        master = adapter.discover_master(gtin)
        click.echo(f"master     {master.master_id}  ({master.product_name})")
        click.echo(f"  pdp      {master.pdp_url}" + ("  [cache]" if master.from_cache else ""))
        click.echo(f"  shades   {len(master.shade_by_gtin)} in swatch list")
        click.echo(f"  size     {master.size_text!r}")
        if master.inci_text:
            shown = master.inci_text if verbose else master.inci_text[:120] + "…"
            click.echo(
                f"  inci     {shown} (captured with shade {master.inci_selected_gtin} selected)"
            )
        if verbose:
            for g, shade in sorted(master.shade_by_gtin.items()):
                click.echo(f"           {g}  {shade}")

        variant = adapter.resolve_variant(master, gtin)
        click.echo(
            f"\nvariant    {variant.gtin13}  ok={variant.ok}  via={variant.via}"
            + ("  [cache]" if variant.from_cache else "")
        )
        if variant.ok:
            click.echo(f"  name     {variant.product_name}")
            click.echo(f"  shade    {variant.shade}")
            click.echo(f"  size     {variant.size_text!r}")
            click.echo(f"  anchor   returned ID == requested GTIN ({variant.returned_id})")
        else:
            click.echo(f"  REJECTED: {variant.reject_reason}")
        click.echo(f"  url      {variant.url}")
        click.echo(f"  snippet  {variant.snippet}")
    finally:
        fetcher.close()
        playwright.close()


@main.command()
@click.option("--run", "run_dir", required=True, type=click.Path(exists=True))
def report(run_dir: str) -> None:
    """Re-print the report for a cached run (Phase 1)."""
    raise click.ClickException("Phase 1: cached runs are not implemented yet")


if __name__ == "__main__":
    main()
